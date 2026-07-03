#!/usr/bin/env python3
"""plot_heine_phi_psi.py
Joint plot of composition x viability (phi . psi) from the Heine workbook
(data/heine_species_distribution_biofilm.xlsx).

For every (species, condition, timepoint, replicate) the workbook gives two
CLSM-measured relative compositions:

  phi   = species share of *all* cells        (sheet block "... all cells")
  L     = species share of *living* cells      ("... only living cells")

so the living-cell composition is the product

  L = phi . psi ,   with   psi = L / phi   (relative viability / enrichment).

The figure makes that multiplication visible: a log-log scatter of psi vs phi
per state (Commensal / Dysbiotic), with the constant-product lines phi.psi = c
drawn as parallel diagonals (each diagonal is one living-composition level c%).
Marginal box-and-whisker plots (IQR) sit on the top (phi) and right (psi) of
each panel, one box per species.

  python plot_heine_phi_psi.py        # -> assets/heine_phi_psi_joint.png
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import openpyxl

_HERE = Path(__file__).resolve().parent
_XLSX = _HERE / "data" / "heine_species_distribution_biofilm.xlsx"
_OUT = _HERE / "assets" / "heine_phi_psi_joint.png"

TAGS = [1, 3, 6, 10, 15, 21]
BASES = ["Static", "HOBIC"]
# 5 species in fixed order (early coloniser -> pathogen); Okabe-Ito CVD-safe.
ROLES = ["S. oralis", "A. naeslundii", "Veillonella", "F. nucleatum", "P. gingivalis"]
COLORS = ["#0072B2", "#56B4E9", "#009E73", "#E69F00", "#D55E00"]
SURFACE = "#fcfcfb"
INK, MUTED = "#222", "#888"

# display window (log): phi in %, psi = living/all ratio
PHI_LO, PHI_HI = 0.05, 100.0
PSI_LO, PSI_HI = 0.05, 30.0
PRODUCT_LEVELS = [1, 5, 20, 50]  # phi.psi = c  (living composition, %)


def load(ws):
    """{condition: [[replicate values]*5 species] per tag}."""
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(r for r in rows if r[1] and str(r[1]).startswith("S. oralis"))
    sp = [j for j in range(1, len(hdr)) if hdr[j]]
    width = sp[1] - sp[0]
    d: dict = {}
    cond = None
    for r in rows:
        v = r[0]
        if isinstance(v, str) and "cells" in v:
            cond = v.strip()
            d[cond] = {}
            continue
        if cond and v in TAGS:
            d[cond][v] = [[r[c + k] for k in range(width)] for c in sp]
    return d, width


def pairs_by_species(d, width):
    """Return per-species arrays of (phi, psi) pooled over base/tag/replicate."""
    phi = [[] for _ in range(5)]
    psi = [[] for _ in range(5)]
    for base in BASES:
        A = d[base + " all cells"]
        L = d[base + " only living cells"]
        for tag in TAGS:
            for si in range(5):
                for j in range(width):
                    a = A[tag][si][j]
                    l = L[tag][si][j]
                    if (isinstance(a, (int, float)) and isinstance(l, (int, float))
                            and a > 1e-9 and l >= 0):
                        phi[si].append(float(a))
                        psi[si].append(l / a)
    return [np.array(p) for p in phi], [np.array(p) for p in psi]


def _box(ax, data, pos, color, vert):
    """One coloured IQR box (median line, 1.5*IQR whiskers, outlier dots)."""
    if len(data) == 0:
        return
    bp = ax.boxplot(
        data, positions=[pos],
        orientation=("vertical" if vert else "horizontal"),
        widths=0.55, whis=1.5,
        patch_artist=True, showcaps=False,
        medianprops=dict(color=SURFACE, linewidth=1.6),
        whiskerprops=dict(color=color, linewidth=1.2),
        boxprops=dict(facecolor=color, edgecolor=color, linewidth=0.0),
        flierprops=dict(marker="o", markersize=2.4, markerfacecolor=color,
                        markeredgecolor="none", alpha=0.5),
    )
    return bp


def joint_panel(fig, sub, phi, psi, title):
    gs = sub.subgridspec(2, 2, width_ratios=[4, 1], height_ratios=[1, 4],
                         wspace=0.04, hspace=0.04)
    axm = fig.add_subplot(gs[1, 0])
    axt = fig.add_subplot(gs[0, 0], sharex=axm)
    axr = fig.add_subplot(gs[1, 1], sharey=axm)

    # --- main: log-log scatter + constant-product diagonals -----------------
    axm.set_xscale("log")
    axm.set_yscale("log")
    xx = np.array([PHI_LO, PHI_HI])
    for c in PRODUCT_LEVELS:
        axm.plot(xx, c / xx, color=MUTED, lw=0.8, ls=(0, (4, 3)), zorder=1)
        # label near the top edge where the line enters the window
        xl = c / PSI_HI
        if PHI_LO < xl < PHI_HI:
            axm.text(xl * 1.15, PSI_HI * 0.72, f"phi.psi={c}", color=MUTED,
                     fontsize=6.5, rotation=-45, ha="left", va="top", zorder=1)
    axm.axhline(1.0, color=MUTED, lw=0.9, zorder=2)  # psi=1: living share == all share
    axm.text(PHI_LO * 1.25, 1.08, "psi = 1  (living share = all share)",
             color=MUTED, fontsize=6.5, ha="left", va="bottom", zorder=2)
    for si in range(5):
        axm.scatter(phi[si], psi[si], s=13, color=COLORS[si], alpha=0.72,
                    edgecolor=SURFACE, linewidth=0.4, zorder=3)
    axm.set_xlim(PHI_LO, PHI_HI)
    axm.set_ylim(PSI_LO, PSI_HI)
    axm.set_xlabel("phi  =  composition, all cells (%)", fontsize=9)
    axm.set_ylabel("psi  =  viability  (living / all)", fontsize=9)
    axm.tick_params(labelsize=8)
    for s in ("top", "right"):
        axm.spines[s].set_visible(False)

    # --- top marginal: phi IQR box per species (shares x) -------------------
    for si in range(5):
        _box(axt, phi[si], pos=si, color=COLORS[si], vert=False)
    axt.set_ylim(-0.6, 4.6)
    axt.set_yticks([])
    axt.tick_params(labelbottom=False, length=0)
    for s in ("top", "right", "left"):
        axt.spines[s].set_visible(False)
    axt.set_title(title, fontsize=11, fontweight="bold", pad=6)

    # --- right marginal: psi IQR box per species (shares y) -----------------
    for si in range(5):
        _box(axr, psi[si], pos=si, color=COLORS[si], vert=True)
    axr.set_xlim(-0.6, 4.6)
    axr.set_xticks([])
    axr.tick_params(labelleft=False, length=0)
    for s in ("top", "right", "bottom"):
        axr.spines[s].set_visible(False)


def main():
    wb = openpyxl.load_workbook(_XLSX, data_only=True, read_only=True)
    states = wb.sheetnames  # ['Commensal', 'Dysbiotic']

    plt.rcParams.update({
        "font.family": "serif",
        "font.serif": ["Times New Roman", "DejaVu Serif"],
        "font.size": 10, "axes.edgecolor": MUTED, "text.color": INK,
        "axes.labelcolor": INK, "xtick.color": INK, "ytick.color": INK,
        "figure.facecolor": SURFACE, "axes.facecolor": SURFACE,
    })
    fig = plt.figure(figsize=(11.6, 5.9))
    outer = fig.add_gridspec(1, len(states), wspace=0.18)
    for i, st in enumerate(states):
        d, width = load(wb[st])
        phi, psi = pairs_by_species(d, width)
        joint_panel(fig, outer[0, i], phi, psi, st)

    handles = [Line2D([0], [0], marker="o", linestyle="none", markersize=7,
                      markerfacecolor=COLORS[k], markeredgecolor="none",
                      label=ROLES[k]) for k in range(5)]
    handles.append(Line2D([0], [0], color=MUTED, lw=0.8, ls=(0, (4, 3)),
                          label="phi.psi = const (living composition %)"))
    fig.legend(handles=handles, loc="lower center", ncol=6, frameon=False,
               fontsize=9, bbox_to_anchor=(0.5, -0.015), handlelength=1.4)
    fig.suptitle(
        "Heine 5-species biofilm  —  composition x viability   (L = phi . psi)",
        fontsize=12.5, fontweight="bold", y=0.99)
    fig.tight_layout(rect=(0, 0.05, 1, 0.96))
    _OUT.parent.mkdir(exist_ok=True)
    fig.savefig(_OUT, dpi=200, bbox_inches="tight")
    print(f"wrote {_OUT}")


if __name__ == "__main__":
    main()
