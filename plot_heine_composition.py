#!/usr/bin/env python3
"""plot_heine_composition.py
Stacked-bar composition figure from the Heine species-distribution workbook
(data/heine_species_distribution_biofilm.xlsx).

Per state (Commensal / Dysbiotic) x condition (Static/HOBIC x all/living cells),
show the mean relative composition of the 5 species over timepoints
1,3,6,10,15,21 d (replicate columns averaged, then normalised to 100%).

  python plot_heine_composition.py            # -> assets/heine_species_composition.png
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

_HERE = Path(__file__).resolve().parent
_XLSX = _HERE / "data" / "heine_species_distribution_biofilm.xlsx"
_OUT = _HERE / "assets" / "heine_species_composition.png"

TAGS = [1, 3, 6, 10, 15, 21]
# 5 species roles in fixed stacking order (early coloniser -> pathogen).
# CVD-safe Okabe-Ito subset, validated (dataviz validator, ΔE>=18, light).
ROLES = ["S. oralis", "A. naeslundii", "Veillonella", "F. nucleatum", "P. gingivalis"]
COLORS = ["#0072B2", "#56B4E9", "#009E73", "#E69F00", "#D55E00"]
SURFACE = "#fcfcfb"


def parse_sheet(ws):
    """Return {condition: ndarray(len(TAGS), 5)} of normalised composition (%)."""
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(r for r in rows if r[1] and str(r[1]).startswith("S. oralis"))
    sp_cols = [j for j in range(1, len(hdr)) if hdr[j]]
    width = sp_cols[1] - sp_cols[0]

    out: dict[str, np.ndarray] = {}
    cond = None
    for r in rows:
        v0 = r[0]
        if isinstance(v0, str) and "cells" in v0:
            cond = v0.strip()
            out[cond] = np.full((len(TAGS), 5), np.nan)
            continue
        if cond is None or v0 not in TAGS:
            continue
        ti = TAGS.index(v0)
        means = []
        for c in sp_cols:
            vals = [r[c + k] for k in range(width)
                    if isinstance(r[c + k], (int, float))]
            means.append(np.mean(vals) if vals else np.nan)
        means = np.array(means, float)
        s = np.nansum(means)
        out[cond][ti] = 100.0 * means / s if s > 0 else np.nan
    return out


def main():
    wb = openpyxl.load_workbook(_XLSX, data_only=True, read_only=True)
    states = wb.sheetnames  # ['Commensal', 'Dysbiotic']
    data = {st: parse_sheet(wb[st]) for st in states}
    conds = list(data[states[0]].keys())  # 4 conditions, in-sheet order

    plt.rcParams.update({
        "font.family": "serif",
        "font.serif": ["Times New Roman", "DejaVu Serif"],
        "font.size": 10, "axes.edgecolor": "#888", "figure.facecolor": SURFACE,
        "axes.facecolor": SURFACE,
    })
    nR, nC = len(states), len(conds)
    fig, axes = plt.subplots(nR, nC, figsize=(3.3 * nC, 3.1 * nR),
                             sharex=True, sharey=True)
    x = np.arange(len(TAGS))
    for i, st in enumerate(states):
        for j, cond in enumerate(conds):
            ax = axes[i, j]
            comp = data[st][cond]                 # (6,5)
            bottom = np.zeros(len(TAGS))
            for k in range(5):
                vals = np.nan_to_num(comp[:, k])
                ax.bar(x, vals, bottom=bottom, width=0.72, color=COLORS[k],
                       edgecolor=SURFACE, linewidth=1.4,
                       label=ROLES[k] if (i == 0 and j == 0) else None)
                # selective direct labels: only clearly dominant segments
                for xi, (v, b) in enumerate(zip(vals, bottom)):
                    if v >= 12:
                        ax.text(xi, b + v / 2, f"{v:.0f}", ha="center",
                                va="center", fontsize=6.5,
                                color="white" if k != 3 else "#333")
                bottom += vals
            ax.set_ylim(0, 100)
            ax.set_xticks(x)
            ax.set_xticklabels([f"{t}" for t in TAGS])
            ax.tick_params(length=0)
            for s in ("top", "right"):
                ax.spines[s].set_visible(False)
            if i == 0:
                ax.set_title(cond, fontsize=9.5, fontweight="bold")
            if j == 0:
                ax.set_ylabel(f"{st}\ncomposition (%)", fontsize=9.5)
            if i == nR - 1:
                ax.set_xlabel("day", fontsize=9)

    handles, labels = axes[0, 0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", ncol=5, frameon=False,
               bbox_to_anchor=(0.5, -0.02), fontsize=9.5, handlelength=1.2)
    fig.suptitle("Heine 5-species biofilm — mean CLSM composition over time",
                 fontsize=12, fontweight="bold", y=0.99)
    fig.tight_layout(rect=(0, 0.04, 1, 0.97))
    _OUT.parent.mkdir(exist_ok=True)
    fig.savefig(_OUT, dpi=200, bbox_inches="tight")
    print(f"wrote {_OUT}")


if __name__ == "__main__":
    main()
